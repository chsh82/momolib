# -*- coding: utf-8 -*-
from flask import render_template, request, redirect, url_for, flash, jsonify, abort
from flask_login import login_required, current_user
from app.library import library_bp
from app.models import db
from app.models.library import (Book, LearningContent, QuizQuestion,
                                  ReadingRecord, ContentCompletion, EssaySubmission,
                                  GENRE_CHOICES, LEVEL_CHOICES, CONTENT_TYPES)
from app.models.reading_mbti import (ReadingMBTITest, ReadingMBTIQuestion,
                                      ReadingMBTIType, ReadingMBTIResponse, ReadingMBTIResult)
from app.models.book_mbti import BookMBTIResult, BOOK_MBTI_TYPES, BOOK_MBTI_QUESTIONS
from app.models.user import User
from app.utils.mileage import award_mileage
from app.models.avatar import MileageReason


# ─────────────────────────────────────────────
# HQ / Admin: 도서 관리
# ─────────────────────────────────────────────

@library_bp.route('/books')
@login_required
def book_list():
    if not (current_user.is_hq or current_user.is_branch_owner or current_user.is_branch_staff):
        abort(403)
    q = request.args.get('q', '').strip()
    genre = request.args.get('genre', '')
    level = request.args.get('level', '')
    query = Book.query.filter_by(is_active=True)
    if q:
        query = query.filter(Book.title.ilike(f'%{q}%') | Book.author.ilike(f'%{q}%'))
    if genre:
        query = query.filter_by(genre=genre)
    if level:
        query = query.filter_by(level=level)
    books = query.order_by(Book.created_at.desc()).all()
    return render_template('library/book_list.html', books=books, q=q,
                           genre=genre, level=level,
                           genre_choices=GENRE_CHOICES, level_choices=LEVEL_CHOICES)


@library_bp.route('/books/new', methods=['GET', 'POST'])
@login_required
def book_new():
    if not current_user.is_hq:
        abort(403)
    if request.method == 'POST':
        book = Book(
            isbn=request.form.get('isbn') or None,
            title=request.form['title'],
            author=request.form.get('author'),
            publisher=request.form.get('publisher'),
            publication_year=request.form.get('publication_year') or None,
            cover_image_url=request.form.get('cover_image_url'),
            description=request.form.get('description'),
            genre=request.form.get('genre'),
            level=request.form.get('level', 'all'),
            tags=request.form.get('tags'),
            created_by=current_user.user_id,
        )
        db.session.add(book)
        db.session.commit()
        flash('도서가 등록되었습니다.', 'success')
        return redirect(url_for('library.book_detail', book_id=book.book_id))
    return render_template('library/book_form.html',
                           genre_choices=GENRE_CHOICES, level_choices=LEVEL_CHOICES)


@library_bp.route('/books/template')
@login_required
def book_template():
    """도서 엑셀 양식 다운로드"""
    if not current_user.is_hq:
        abort(403)
    import openpyxl
    from flask import send_file
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '도서목록'

    headers = ['제목*', '저자', '출판사', '출판연도', 'ISBN', '장르', '수준', '쪽수', '태그(쉼표구분)', '표지URL', '설명']
    ws.append(headers)

    # 헤더 스타일
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='solid', fgColor='E8EAF6')

    # 장르/수준 안내 시트
    ws2 = wb.create_sheet('장르코드')
    ws2.append(['코드', '장르명'])
    for val, label in GENRE_CHOICES:
        ws2.append([val, label])

    ws3 = wb.create_sheet('수준코드')
    ws3.append(['코드', '수준명'])
    for val, label in LEVEL_CHOICES:
        ws3.append([val, label])

    # 예시 행
    ws.append(['파친코', '이민진', '인플루엔셜', 2017, '9791186560846',
               'literature', 'high', 490, '소설,디아스포라', '', '재일 한국인 4대에 걸친 이야기'])

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='도서등록양식.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@library_bp.route('/books/bulk-upload', methods=['POST'])
@login_required
def book_bulk_upload():
    """도서 엑셀 일괄 업로드"""
    if not current_user.is_hq:
        abort(403)
    import openpyxl

    f = request.files.get('excel_file')
    if not f:
        flash('파일을 선택해주세요.', 'error')
        return redirect(url_for('library.book_list'))

    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception:
        flash('엑셀 파일을 읽을 수 없습니다.', 'error')
        return redirect(url_for('library.book_list'))

    added = skipped = errors = 0
    error_msgs = []

    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        title = str(row[0]).strip()
        if not title:
            continue

        author        = str(row[1]).strip() if row[1] else None
        publisher     = str(row[2]).strip() if row[2] else None
        pub_year      = int(row[3]) if row[3] and str(row[3]).strip().isdigit() else None
        isbn          = str(row[4]).strip() if row[4] else None
        genre         = str(row[5]).strip() if row[5] else None
        level         = str(row[6]).strip() if row[6] else 'all'
        page_count    = int(row[7]) if row[7] and str(row[7]).strip().isdigit() else None
        tags          = str(row[8]).strip() if row[8] else None
        cover_url     = str(row[9]).strip() if row[9] else None
        description   = str(row[10]).strip() if row[10] else None

        # ISBN 중복 체크
        if isbn:
            exists = Book.query.filter_by(isbn=isbn).first()
            if exists:
                skipped += 1
                continue

        # 장르/수준 유효성
        valid_genres = [v for v, _ in GENRE_CHOICES]
        valid_levels = [v for v, _ in LEVEL_CHOICES]
        if genre and genre not in valid_genres:
            error_msgs.append(f'{i}행: 장르 코드 "{genre}" 오류')
            errors += 1
            continue
        if level not in valid_levels:
            level = 'all'

        try:
            book = Book(
                title=title,
                author=author,
                publisher=publisher,
                publication_year=pub_year,
                isbn=isbn or None,
                genre=genre or None,
                level=level,
                page_count=page_count,
                tags=tags,
                cover_image_url=cover_url,
                description=description,
                created_by=current_user.user_id,
            )
            db.session.add(book)
            added += 1
        except Exception as e:
            error_msgs.append(f'{i}행 오류: {str(e)}')
            errors += 1

    db.session.commit()

    msg = f'{added}권 등록 완료.'
    if skipped:
        msg += f' {skipped}권 중복(ISBN) 건너뜀.'
    if errors:
        msg += f' {errors}건 오류.'
    flash(msg, 'success' if not errors else 'warning')

    if error_msgs:
        for m in error_msgs[:5]:
            flash(m, 'error')

    return redirect(url_for('library.book_list'))


@library_bp.route('/books/isbn-lookup')
@login_required
def isbn_lookup():
    """ISBN AJAX 조회"""
    isbn = request.args.get('isbn', '').strip()
    if not isbn:
        return jsonify({'error': 'ISBN을 입력하세요.'})
    try:
        from app.services.isbn_service import ISBNService
        svc = ISBNService()
        data = svc.lookup_isbn(isbn)
        if data:
            return jsonify({'ok': True, 'data': data})
        return jsonify({'ok': False, 'error': '검색 결과가 없습니다.'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@library_bp.route('/books/<book_id>')
@login_required
def book_detail(book_id):
    book = Book.query.get_or_404(book_id)
    return render_template('library/book_detail.html', book=book,
                           content_types=CONTENT_TYPES)


@library_bp.route('/books/<book_id>/edit', methods=['GET', 'POST'])
@login_required
def book_edit(book_id):
    if not current_user.is_hq:
        abort(403)
    book = Book.query.get_or_404(book_id)
    if request.method == 'POST':
        book.title = request.form['title']
        book.author = request.form.get('author')
        book.publisher = request.form.get('publisher')
        book.publication_year = request.form.get('publication_year') or None
        book.cover_image_url = request.form.get('cover_image_url')
        book.description = request.form.get('description')
        book.genre = request.form.get('genre')
        book.level = request.form.get('level', 'all')
        book.tags = request.form.get('tags')
        db.session.commit()
        flash('수정되었습니다.', 'success')
        return redirect(url_for('library.book_detail', book_id=book_id))
    return render_template('library/book_form.html', book=book,
                           genre_choices=GENRE_CHOICES, level_choices=LEVEL_CHOICES)


@library_bp.route('/books/<book_id>/delete', methods=['POST'])
@login_required
def book_delete(book_id):
    if not current_user.is_hq:
        abort(403)
    book = Book.query.get_or_404(book_id)
    book.is_active = False
    db.session.commit()
    flash('도서가 삭제되었습니다.', 'success')
    return redirect(url_for('library.book_list'))


# ─────────────────────────────────────────────
# HQ: 콘텐츠 관리
# ─────────────────────────────────────────────

@library_bp.route('/books/<book_id>/contents/new', methods=['GET', 'POST'])
@login_required
def content_new(book_id):
    if not current_user.is_hq:
        abort(403)
    book = Book.query.get_or_404(book_id)
    if request.method == 'POST':
        ctype = request.form['type']
        data = {}
        if ctype == 'video':
            data = {
                'url': request.form.get('video_url', ''),
                'duration_seconds': int(request.form.get('duration_seconds') or 0),
                'thumbnail_url': request.form.get('thumbnail_url', ''),
            }
        elif ctype == 'essay':
            data = {
                'prompt': request.form.get('prompt', ''),
                'rubric': request.form.get('rubric', ''),
                'max_score': float(request.form.get('max_score') or 100),
            }
        content = LearningContent(
            book_id=book_id,
            type=ctype,
            title=request.form['title'],
            description=request.form.get('description'),
            data=data,
            order_num=int(request.form.get('order_num') or 0),
            is_published=bool(request.form.get('is_published')),
            created_by=current_user.user_id,
        )
        db.session.add(content)
        db.session.commit()
        flash('콘텐츠가 추가되었습니다.', 'success')
        return redirect(url_for('library.book_detail', book_id=book_id))
    return render_template('library/content_form.html', book=book,
                           content_types=CONTENT_TYPES)


@library_bp.route('/contents/<content_id>/edit', methods=['GET', 'POST'])
@login_required
def content_edit(content_id):
    if not current_user.is_hq:
        abort(403)
    content = LearningContent.query.get_or_404(content_id)
    if request.method == 'POST':
        content.title = request.form['title']
        content.description = request.form.get('description')
        content.order_num = int(request.form.get('order_num') or 0)
        content.is_published = bool(request.form.get('is_published'))
        if content.type == 'video':
            content.data = {
                'url': request.form.get('video_url', ''),
                'duration_seconds': int(request.form.get('duration_seconds') or 0),
                'thumbnail_url': request.form.get('thumbnail_url', ''),
            }
        elif content.type == 'essay':
            content.data = {
                'prompt': request.form.get('prompt', ''),
                'rubric': request.form.get('rubric', ''),
                'max_score': float(request.form.get('max_score') or 100),
            }
        db.session.commit()
        flash('수정되었습니다.', 'success')
        return redirect(url_for('library.book_detail', book_id=content.book_id))
    return render_template('library/content_form.html', book=content.book,
                           content=content, content_types=CONTENT_TYPES)


@library_bp.route('/contents/<content_id>/delete', methods=['POST'])
@login_required
def content_delete(content_id):
    if not current_user.is_hq:
        abort(403)
    content = LearningContent.query.get_or_404(content_id)
    book_id = content.book_id
    db.session.delete(content)
    db.session.commit()
    flash('콘텐츠가 삭제되었습니다.', 'success')
    return redirect(url_for('library.book_detail', book_id=book_id))


# ─────────────────────────────────────────────
# HQ: 퀴즈 문항 관리
# ─────────────────────────────────────────────

@library_bp.route('/contents/<content_id>/questions')
@login_required
def question_list(content_id):
    if not current_user.is_hq:
        abort(403)
    content = LearningContent.query.get_or_404(content_id)
    return render_template('library/question_list.html', content=content)


@library_bp.route('/contents/<content_id>/questions/new', methods=['GET', 'POST'])
@login_required
def question_new(content_id):
    if not current_user.is_hq:
        abort(403)
    content = LearningContent.query.get_or_404(content_id)
    if request.method == 'POST':
        q = QuizQuestion(
            content_id=content_id,
            question_text=request.form['question_text'],
            choices=_parse_choices(request.form, content.type),
            correct_answer=request.form.get('correct_answer'),
            explanation=request.form.get('explanation'),
            order_num=int(request.form.get('order_num') or 0),
        )
        db.session.add(q)
        db.session.commit()
        flash('문항이 추가되었습니다.', 'success')
        return redirect(url_for('library.question_list', content_id=content_id))
    return render_template('library/question_form.html', content=content)


@library_bp.route('/questions/<int:question_id>/edit', methods=['GET', 'POST'])
@login_required
def question_edit(question_id):
    if not current_user.is_hq:
        abort(403)
    q = QuizQuestion.query.get_or_404(question_id)
    if request.method == 'POST':
        q.question_text = request.form['question_text']
        q.choices = _parse_choices(request.form, q.content.type)
        q.correct_answer = request.form.get('correct_answer')
        q.explanation = request.form.get('explanation')
        q.order_num = int(request.form.get('order_num') or 0)
        db.session.commit()
        flash('수정되었습니다.', 'success')
        return redirect(url_for('library.question_list', content_id=q.content_id))
    return render_template('library/question_form.html',
                           content=q.content, question=q)


@library_bp.route('/questions/<int:question_id>/delete', methods=['POST'])
@login_required
def question_delete(question_id):
    if not current_user.is_hq:
        abort(403)
    q = QuizQuestion.query.get_or_404(question_id)
    content_id = q.content_id
    db.session.delete(q)
    db.session.commit()
    return redirect(url_for('library.question_list', content_id=content_id))


def _parse_choices(form, content_type):
    if content_type == 'quiz':
        choices = []
        i = 0
        while f'choice_{i}' in form:
            choices.append({
                'text': form[f'choice_{i}'],
                'is_correct': f'correct_{i}' in form,
            })
            i += 1
        return choices
    elif content_type == 'initial_quiz':
        return {'answer': form.get('initial_answer', ''), 'hint': form.get('hint', '')}
    elif content_type == 'vocab_quiz':
        choices = []
        i = 0
        while f'choice_{i}' in form:
            choices.append(form[f'choice_{i}'])
            i += 1
        return {
            'word': form.get('word', ''),
            'choices': choices,
            'correct_idx': int(form.get('correct_idx') or 0),
        }
    return None


# ─────────────────────────────────────────────
# HQ: 서술형 채점
# ─────────────────────────────────────────────

@library_bp.route('/essay-submissions')
@login_required
def essay_submissions():
    if not (current_user.is_hq or current_user.is_branch_owner or current_user.is_branch_staff):
        abort(403)
    pending = request.args.get('pending', '')
    query = EssaySubmission.query
    if pending:
        query = query.filter(EssaySubmission.score.is_(None))
    submissions = query.order_by(EssaySubmission.submitted_at.desc()).limit(100).all()
    return render_template('library/essay_submissions.html',
                           submissions=submissions, pending=pending)


@library_bp.route('/essay-submissions/<submission_id>', methods=['GET', 'POST'])
@login_required
def essay_submission_grade(submission_id):
    if not (current_user.is_hq or current_user.is_branch_owner or current_user.is_branch_staff):
        abort(403)
    sub = EssaySubmission.query.get_or_404(submission_id)
    if request.method == 'POST':
        sub.score = float(request.form['score'])
        sub.max_score = float(request.form.get('max_score') or sub.content.data.get('max_score', 100))
        sub.feedback = request.form.get('feedback')
        sub.graded_by = current_user.user_id
        from datetime import datetime
        sub.graded_at = datetime.utcnow()
        db.session.commit()
        flash('채점이 완료되었습니다.', 'success')
        return redirect(url_for('library.essay_submissions', pending=1))
    return render_template('library/essay_submission_grade.html', sub=sub)


# ─────────────────────────────────────────────
# Student: 도서 탐색 & 학습
# ─────────────────────────────────────────────

@library_bp.route('/catalog')
@login_required
def catalog():
    q = request.args.get('q', '').strip()
    genre = request.args.get('genre', '')
    level = request.args.get('level', '')
    query = Book.query.filter_by(is_active=True)
    if q:
        query = query.filter(Book.title.ilike(f'%{q}%') | Book.author.ilike(f'%{q}%'))
    if genre:
        query = query.filter_by(genre=genre)
    if level:
        query = query.filter_by(level=level)
    books = query.order_by(Book.created_at.desc()).all()

    # 이미 등록된 독서 기록
    my_records = {r.book_id: r for r in
                  ReadingRecord.query.filter_by(student_id=current_user.user_id).all()}

    return render_template('library/catalog.html', books=books, q=q,
                           genre=genre, level=level,
                           genre_choices=GENRE_CHOICES, level_choices=LEVEL_CHOICES,
                           my_records=my_records)


@library_bp.route('/catalog/<book_id>/start', methods=['POST'])
@login_required
def start_reading(book_id):
    book = Book.query.get_or_404(book_id)
    existing = ReadingRecord.query.filter_by(
        student_id=current_user.user_id, book_id=book_id).first()
    if not existing:
        record = ReadingRecord(
            student_id=current_user.user_id,
            book_id=book_id,
            branch_id=current_user.branch_id,
        )
        db.session.add(record)
        db.session.commit()
        flash(f'"{book.title}" 독서를 시작했습니다!', 'success')
    return redirect(url_for('library.my_book', book_id=book_id))


@library_bp.route('/my-books')
@login_required
def my_books():
    records = (ReadingRecord.query
               .filter_by(student_id=current_user.user_id)
               .order_by(ReadingRecord.started_at.desc()).all())
    return render_template('library/my_books.html', records=records)


@library_bp.route('/my-books/<book_id>')
@login_required
def my_book(book_id):
    book = Book.query.get_or_404(book_id)
    record = ReadingRecord.query.filter_by(
        student_id=current_user.user_id, book_id=book_id).first()
    if not record:
        return redirect(url_for('library.catalog'))

    # 완료된 콘텐츠 목록
    completed_ids = {c.content_id for c in
                     ContentCompletion.query.filter_by(student_id=current_user.user_id).all()}

    contents = [c for c in book.contents if c.is_published]
    return render_template('library/my_book.html', book=book, record=record,
                           contents=contents, completed_ids=completed_ids)


@library_bp.route('/my-books/<book_id>/finish', methods=['POST'])
@login_required
def finish_reading(book_id):
    record = ReadingRecord.query.filter_by(
        student_id=current_user.user_id, book_id=book_id).first_or_404()
    from datetime import datetime
    record.status = 'completed'
    record.finished_at = datetime.utcnow()
    record.rating = request.form.get('rating') or None
    record.review = request.form.get('review')
    award_mileage(current_user.user_id, MileageReason.BOOK_FINISH,
                  description=f'독서 완료: {record.book.title}',
                  ref_type='book_finish', ref_id=book_id)
    db.session.commit()
    flash('독서 완료를 기록했습니다!', 'success')
    return redirect(url_for('library.my_books'))


@library_bp.route('/contents/<content_id>/play')
@login_required
def content_play(content_id):
    content = LearningContent.query.get_or_404(content_id)
    if not content.is_published:
        abort(404)
    record = ReadingRecord.query.filter_by(
        student_id=current_user.user_id, book_id=content.book_id).first()
    if not record:
        abort(403)
    completion = ContentCompletion.query.filter_by(
        student_id=current_user.user_id, content_id=content_id).first()
    return render_template('library/content_play.html',
                           content=content, completion=completion)


@library_bp.route('/contents/<content_id>/submit', methods=['POST'])
@login_required
def content_submit(content_id):
    content = LearningContent.query.get_or_404(content_id)
    ctype = content.type

    if ctype == 'video':
        _complete_content(content_id, score=None, max_score=None, answer_data=None)
        award_mileage(current_user.user_id, MileageReason.LIBRARY_CONTENT,
                      description=f'도서 영상: {content.title}',
                      ref_type='library_content', ref_id=content_id)
        flash('영상을 완료했습니다.', 'success')

    elif ctype in ('quiz', 'initial_quiz', 'vocab_quiz'):
        answers = {}
        score = 0
        max_score = len(content.questions)
        for q in content.questions:
            ans = request.form.get(f'q_{q.question_id}')
            answers[str(q.question_id)] = ans
            if _check_answer(q, ans, ctype):
                score += 1
        _complete_content(content_id, score=score, max_score=max_score,
                          answer_data=answers)
        award_mileage(current_user.user_id, MileageReason.LIBRARY_CONTENT,
                      description=f'도서 퀴즈: {content.title}',
                      ref_type='library_content', ref_id=content_id)
        flash(f'퀴즈 완료! {score}/{max_score}점', 'success')

    elif ctype == 'essay':
        text = request.form.get('essay_text', '').strip()
        if not text:
            flash('내용을 입력해주세요.', 'warning')
            return redirect(url_for('library.content_play', content_id=content_id))
        # 이미 제출한 경우 중복 방지
        existing = EssaySubmission.query.filter_by(
            student_id=current_user.user_id, content_id=content_id).first()
        if not existing:
            sub = EssaySubmission(
                student_id=current_user.user_id,
                content_id=content_id,
                text=text,
                max_score=float((content.data or {}).get('max_score', 100)),
            )
            db.session.add(sub)
            db.session.commit()
        _complete_content(content_id, score=None, max_score=None, answer_data=None)
        flash('서술형 평가를 제출했습니다. 채점 후 점수가 반영됩니다.', 'success')

    return redirect(url_for('library.my_book', book_id=content.book_id))


def _check_answer(question, answer, ctype):
    if answer is None:
        return False
    if ctype == 'quiz':
        return str(question.correct_answer) == str(answer)
    elif ctype == 'initial_quiz':
        return (question.choices or {}).get('answer', '').strip() == answer.strip()
    elif ctype == 'vocab_quiz':
        return str((question.choices or {}).get('correct_idx', -1)) == str(answer)
    return False


def _complete_content(content_id, score, max_score, answer_data):
    existing = ContentCompletion.query.filter_by(
        student_id=current_user.user_id, content_id=content_id).first()
    if existing:
        if score is not None:
            existing.score = score
            existing.max_score = max_score
            existing.answer_data = answer_data
    else:
        c = ContentCompletion(
            student_id=current_user.user_id,
            content_id=content_id,
            score=score,
            max_score=max_score,
            answer_data=answer_data,
        )
        db.session.add(c)
    db.session.commit()


# ─────────────────────────────────────────────
# Student: 학습 분석
# ─────────────────────────────────────────────

@library_bp.route('/my-analytics')
@login_required
def my_analytics():
    records = ReadingRecord.query.filter_by(student_id=current_user.user_id).all()
    completions = ContentCompletion.query.filter_by(student_id=current_user.user_id).all()

    # 장르별 독서수
    genre_counts = {}
    for r in records:
        g = r.book.genre_display if r.book else '기타'
        genre_counts[g] = genre_counts.get(g, 0) + 1

    # 콘텐츠 유형별 완료수
    type_counts = {}
    for c in completions:
        t = c.content.type_display if c.content else '기타'
        type_counts[t] = type_counts.get(t, 0) + 1

    # 평균 점수 (퀴즈류만)
    scored = [c for c in completions if c.score is not None and c.max_score]
    avg_pct = round(sum(c.score_pct for c in scored) / len(scored)) if scored else None

    mbti_result = (ReadingMBTIResult.query
                   .filter_by(user_id=current_user.user_id)
                   .order_by(ReadingMBTIResult.submitted_at.desc()).first())

    # LMS 학습 데이터
    from app.models.lms import StudentPackageAssignment, StudentItemProgress
    lms_assignments = StudentPackageAssignment.query.filter_by(
        student_id=current_user.user_id, is_active=True
    ).all()

    lms_progresses = StudentItemProgress.query.filter_by(
        student_id=current_user.user_id, status='completed'
    ).all()

    # 패키지별 진도
    lms_package_stats = []
    for a in lms_assignments:
        total = sum(len(pc.curriculum.items) for pc in a.package.curricula)
        done = StudentItemProgress.query.filter_by(
            student_id=current_user.user_id,
            assignment_id=a.id,
            status='completed'
        ).count()
        pct = round(done / total * 100) if total > 0 else 0
        lms_package_stats.append({
            'title': a.package.title,
            'done': done,
            'total': total,
            'pct': pct,
            'completed': done == total and total > 0,
        })

    # LMS 콘텐츠 유형별 완료수
    lms_type_counts = {}
    for p in lms_progresses:
        t = p.item.content_type_display if p.item else '기타'
        lms_type_counts[t] = lms_type_counts.get(t, 0) + 1

    # LMS 평균 퀴즈 점수
    lms_scored = [p for p in lms_progresses if p.score is not None]
    lms_avg_pct = round(sum(p.score for p in lms_scored) / len(lms_scored) * 100) if lms_scored else None

    return render_template('library/my_analytics.html',
                           records=records, completions=completions,
                           genre_counts=genre_counts, type_counts=type_counts,
                           avg_pct=avg_pct, mbti_result=mbti_result,
                           lms_package_stats=lms_package_stats,
                           lms_type_counts=lms_type_counts,
                           lms_avg_pct=lms_avg_pct,
                           lms_total_done=len(lms_progresses))


# ─────────────────────────────────────────────
# Parent: 자녀 학습 현황
# ─────────────────────────────────────────────

@library_bp.route('/parent/child/<child_id>')
@login_required
def parent_child_analytics(child_id):
    if current_user.role != 'parent':
        abort(403)
    # 자녀 연결 확인
    from app.models.member import ParentStudent
    link = ParentStudent.query.filter_by(
        parent_id=current_user.user_id, student_id=child_id).first_or_404()
    child = User.query.get_or_404(child_id)

    records = ReadingRecord.query.filter_by(student_id=child_id).all()
    completions = ContentCompletion.query.filter_by(student_id=child_id).all()

    genre_counts = {}
    for r in records:
        g = r.book.genre_display if r.book else '기타'
        genre_counts[g] = genre_counts.get(g, 0) + 1

    scored = [c for c in completions if c.score is not None and c.max_score]
    avg_pct = round(sum(c.score_pct for c in scored) / len(scored)) if scored else None

    mbti_result = (ReadingMBTIResult.query
                   .filter_by(user_id=child_id)
                   .order_by(ReadingMBTIResult.submitted_at.desc()).first())

    return render_template('library/parent_child_analytics.html',
                           child=child, records=records, completions=completions,
                           genre_counts=genre_counts, avg_pct=avg_pct,
                           mbti_result=mbti_result)


# ─────────────────────────────────────────────
# 독서MBTI
# ─────────────────────────────────────────────

@library_bp.route('/mbti')
@login_required
def mbti_intro():
    test = ReadingMBTITest.query.filter_by(is_active=True).first()
    my_result = None
    all_results = []
    if current_user.role in ('student',):
        my_result = (ReadingMBTIResult.query
                     .filter_by(user_id=current_user.user_id)
                     .order_by(ReadingMBTIResult.submitted_at.desc()).first())
        all_results = (ReadingMBTIResult.query
                       .filter_by(user_id=current_user.user_id)
                       .order_by(ReadingMBTIResult.submitted_at.desc()).all())
    return render_template('library/mbti_intro.html', test=test,
                           my_result=my_result, all_results=all_results)


@library_bp.route('/mbti/test')
@login_required
def mbti_test():
    test = ReadingMBTITest.query.filter_by(is_active=True).first()
    if not test:
        flash('테스트가 준비되지 않았습니다.', 'warning')
        return redirect(url_for('library.mbti_intro'))
    all_questions = test.questions
    absolute_questions = [q for q in all_questions if q.question_type == 'absolute']
    comparison_questions = [q for q in all_questions if q.question_type == 'comparison']
    return render_template('library/mbti_test.html', test=test,
                           absolute_questions=absolute_questions,
                           comparison_questions=comparison_questions)


@library_bp.route('/mbti/submit', methods=['POST'])
@login_required
def mbti_submit():
    test = ReadingMBTITest.query.filter_by(is_active=True).first()
    if not test:
        abort(404)

    # 응답 수집
    responses = {}
    for i in range(1, 46):
        val = request.form.get(f'q{i}')
        if val:
            responses[f'q{i}'] = val
    for i in range(1, 6):
        val = request.form.get(f'comp{i}')
        if val:
            responses[f'comp{i}'] = val

    # 유효성 검사
    from app.utils.mbti_calculator import calculate_mbti_scores, determine_mbti_type, validate_responses
    is_valid, err_msg = validate_responses(responses)
    if not is_valid:
        flash('모든 문항에 응답해주세요.', 'warning')
        return redirect(url_for('library.mbti_test'))

    # 점수 계산
    scores = calculate_mbti_scores(responses)
    reading_level, thinking_level, writing_level, type_code = determine_mbti_type(scores)

    result = ReadingMBTIResult(
        user_id=current_user.user_id,
        test_id=test.test_id,
        reading_score=sum(scores['reading'].values()),
        thinking_score=sum(scores['thinking'].values()),
        writing_score=sum(scores['writing'].values()),
        reading_level=reading_level,
        thinking_level=thinking_level,
        writing_level=writing_level,
        type_code=type_code,
        scores=scores,
    )
    db.session.add(result)
    award_mileage(current_user.user_id, MileageReason.MBTI_TEST,
                  description='독서MBTI 검사 완료',
                  ref_type='mbti_result', ref_id=result.result_id)
    db.session.commit()
    return redirect(url_for('library.mbti_result', result_id=result.result_id))


@library_bp.route('/mbti/result/<result_id>')
@login_required
def mbti_result(result_id):
    result = ReadingMBTIResult.query.get_or_404(result_id)
    if result.user_id != current_user.user_id and not current_user.is_hq:
        abort(403)
    type_info = ReadingMBTIType.query.filter_by(type_code=result.type_code).first()
    return render_template('library/mbti_result.html',
                           result=result, type_info=type_info)


# ── 독서MBTI (취향 테스트) ────────────────────────────────

@library_bp.route('/book-mbti')
@login_required
def book_mbti_intro():
    last = BookMBTIResult.query.filter_by(user_id=current_user.user_id)\
        .order_by(BookMBTIResult.taken_at.desc()).first()
    return render_template('library/book_mbti_intro.html',
                           last_result=last,
                           types=BOOK_MBTI_TYPES,
                           total=len(BOOK_MBTI_QUESTIONS))


@library_bp.route('/book-mbti/test')
@login_required
def book_mbti_test():
    return render_template('library/book_mbti_test.html',
                           questions=BOOK_MBTI_QUESTIONS,
                           total=len(BOOK_MBTI_QUESTIONS))


@library_bp.route('/book-mbti/submit', methods=['POST'])
@login_required
def book_mbti_submit():
    scores = {k: 0 for k in BOOK_MBTI_TYPES}
    for i, q in enumerate(BOOK_MBTI_QUESTIONS):
        ans = request.form.get(f'q{i}', type=int)
        if ans is not None and 0 <= ans < len(q['options']):
            for type_key, pts in q['options'][ans]['scores'].items():
                scores[type_key] = scores.get(type_key, 0) + pts

    import random
    max_score = max(scores.values())
    tied = [k for k, v in scores.items() if v == max_score]
    best_type = random.choice(tied) if len(tied) > 1 else tied[0]

    result = BookMBTIResult(
        user_id=current_user.user_id,
        type_code=best_type,
        scores=scores,
    )
    db.session.add(result)

    award_mileage(
        student_id=current_user.user_id,
        reason=MileageReason.MBTI_TEST,
        description='독서MBTI 테스트 완료',
        ref_type='book_mbti',
        ref_id=result.result_id,
    )
    db.session.commit()
    return redirect(url_for('library.book_mbti_result', result_id=result.result_id))


@library_bp.route('/book-mbti/result/<result_id>')
@login_required
def book_mbti_result(result_id):
    result = BookMBTIResult.query.get_or_404(result_id)
    if result.user_id != current_user.user_id:
        abort(403)
    return render_template('library/book_mbti_result.html', result=result)
