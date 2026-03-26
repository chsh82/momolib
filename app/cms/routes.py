# -*- coding: utf-8 -*-
import os
import uuid
from datetime import datetime
from flask import (render_template, redirect, url_for, flash,
                   request, jsonify, send_from_directory, abort)
from flask_login import login_required, current_user
from app.cms import cms_bp
from app.models import db
from app.models.branch import Branch
from app.models.content import ContentItem, ContentPermission, ContentView
from app.models.library import Book
from app.models.content_bank import (
    BankQuestion, LectureVideo, MockExam, MockExamQuestion, StudyMaterial,
    BANK_QUESTION_TYPES, DIFFICULTY_CHOICES, EXAM_QUESTION_TYPES, MATERIAL_TYPES,
    VOCAB_CATEGORIES, READING_CATEGORIES, READING_TYPE_CHOICES,
)
from app.utils.decorators import requires_role
from app.models.notification import Notification

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), '..', 'static', 'uploads', 'materials')
ALLOWED_EXTENSIONS = {'pdf', 'hwp', 'docx', 'pptx', 'xlsx', 'zip'}


def _ensure_upload_dir():
    os.makedirs(UPLOAD_DIR, exist_ok=True)


def _allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _save_file(file_obj):
    """파일 저장 후 (저장경로, 원본명, 확장자, 크기) 반환"""
    _ensure_upload_dir()
    original_name = file_obj.filename
    ext = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else 'bin'
    saved_name = f"{uuid.uuid4().hex}.{ext}"
    full_path = os.path.join(UPLOAD_DIR, saved_name)
    file_obj.save(full_path)
    size = os.path.getsize(full_path)
    return f"uploads/materials/{saved_name}", original_name, ext, size


def _hq_only():
    return current_user.is_hq


# ═══════════════════════════════════════════════
# CMS 메인 인덱스 (공지)
# ═══════════════════════════════════════════════

@cms_bp.route('/')
@login_required
@requires_role('super_admin', 'hq_manager')
def index():
    items = ContentItem.query.order_by(ContentItem.created_at.desc()).all()
    return render_template('cms/index.html', items=items)


@cms_bp.route('/new', methods=['GET', 'POST'])
@login_required
@requires_role('super_admin', 'hq_manager')
def new_content():
    branches = Branch.query.filter_by(status='active').order_by(Branch.code).all()
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        content_type = request.form.get('content_type', 'notice')
        body = request.form.get('body', '').strip()
        is_global = request.form.get('is_global') == 'on'
        target_branches = request.form.getlist('target_branches')
        is_published = request.form.get('is_published') == 'on'
        if not title:
            flash('제목을 입력해주세요.', 'error')
            return render_template('cms/new_content.html', branches=branches)
        item = ContentItem(title=title, content_type=content_type, body=body,
                           is_global=is_global, is_published=is_published,
                           created_by=current_user.user_id)
        db.session.add(item)
        db.session.flush()
        if not is_global and target_branches:
            for branch_id in target_branches:
                db.session.add(ContentPermission(content_id=item.content_id,
                                                  branch_id=branch_id,
                                                  granted_by=current_user.user_id))
        db.session.flush()
        if is_published:
            notif_link = url_for('branch.notice_detail', content_id=item.content_id)
            if is_global:
                Notification.send_to_all_branches(
                    title=f'[본사 공지] {title}', notif_type='new_notice',
                    link_url=notif_link,
                    roles=['branch_owner', 'branch_manager', 'teacher'])
            elif target_branches:
                for bid in target_branches:
                    Notification.send_to_branch(branch_id=bid,
                                                title=f'[본사 공지] {title}',
                                                notif_type='new_notice',
                                                link_url=notif_link,
                                                roles=['branch_owner', 'branch_manager', 'teacher'])
        db.session.commit()
        flash(f'콘텐츠가 {"발행" if is_published else "저장"}되었습니다.', 'success')
        return redirect(url_for('cms.index'))
    return render_template('cms/new_content.html', branches=branches)


@cms_bp.route('/<content_id>/publish', methods=['POST'])
@login_required
@requires_role('super_admin', 'hq_manager')
def publish(content_id):
    item = ContentItem.query.get_or_404(content_id)
    item.is_published = not item.is_published
    db.session.commit()
    status = '발행' if item.is_published else '발행 취소'
    return jsonify({'success': True, 'published': item.is_published, 'message': f'{status}되었습니다.'})


@cms_bp.route('/<content_id>/views')
@login_required
@requires_role('super_admin', 'hq_manager')
def view_stats(content_id):
    item = ContentItem.query.get_or_404(content_id)
    views = ContentView.query.filter_by(content_id=content_id)\
        .order_by(ContentView.viewed_at.desc()).all()
    branches = Branch.query.filter_by(status='active').all()
    return render_template('cms/view_stats.html', item=item, views=views, branches=branches)


# ═══════════════════════════════════════════════
# 공통 헬퍼
# ═══════════════════════════════════════════════

def _all_books():
    return Book.query.filter_by(is_active=True).order_by(Book.title).all()


# ═══════════════════════════════════════════════
# 1. 어휘 퀴즈 관리
# ═══════════════════════════════════════════════

@cms_bp.route('/vocab')
@login_required
def vocab_list():
    if not _hq_only(): abort(403)
    q = request.args.get('q', '').strip()
    book_id = request.args.get('book_id', '')
    cat_large  = request.args.get('cat_large', '')
    cat_medium = request.args.get('cat_medium', '')
    cat_small  = request.args.get('cat_small', '')

    query = BankQuestion.query.filter_by(type='vocab_quiz', is_active=True)
    if q:
        query = query.filter(BankQuestion.title.ilike(f'%{q}%') |
                             BankQuestion.tags.ilike(f'%{q}%'))
    if book_id:
        query = query.filter_by(book_id=book_id)
    if cat_large:
        query = query.filter_by(cat_large=cat_large)
    if cat_medium:
        query = query.filter_by(cat_medium=cat_medium)
    if cat_small:
        query = query.filter_by(cat_small=cat_small)

    questions = query.order_by(BankQuestion.created_at.desc()).all()
    books = _all_books()
    return render_template('cms/vocab/list.html', questions=questions,
                           q=q, book_id=book_id, books=books,
                           cat_large=cat_large, cat_medium=cat_medium, cat_small=cat_small,
                           vocab_categories=VOCAB_CATEGORIES)


@cms_bp.route('/vocab/new', methods=['GET', 'POST'])
@login_required
def vocab_new():
    if not _hq_only(): abort(403)
    if request.method == 'POST':
        choices = [request.form.get(f'choice_{i}', '') for i in range(4)]
        data = {
            'word': request.form.get('word', ''),
            'definition': request.form.get('definition', ''),
            'choices': choices,
            'correct_idx': int(request.form.get('correct_idx', 0)),
        }
        q = BankQuestion(
            type='vocab_quiz',
            title=request.form['title'],
            book_id=request.form.get('book_id') or None,
            week_num=request.form.get('week_num') or None,
            difficulty=request.form.get('difficulty', 'medium'),
            tags=request.form.get('tags'),
            cat_large=request.form.get('cat_large') or None,
            cat_medium=request.form.get('cat_medium') or None,
            cat_small=request.form.get('cat_small') or None,
            data=data,
            created_by=current_user.user_id,
        )
        db.session.add(q)
        db.session.commit()
        flash('어휘 퀴즈가 등록되었습니다.', 'success')
        return redirect(url_for('cms.vocab_list'))
    return render_template('cms/vocab/form.html', books=_all_books(),
                           difficulty_choices=DIFFICULTY_CHOICES,
                           vocab_categories=VOCAB_CATEGORIES)


@cms_bp.route('/vocab/<question_id>/edit', methods=['GET', 'POST'])
@login_required
def vocab_edit(question_id):
    if not _hq_only(): abort(403)
    q = BankQuestion.query.filter_by(question_id=question_id, type='vocab_quiz').first_or_404()
    if request.method == 'POST':
        choices = [request.form.get(f'choice_{i}', '') for i in range(4)]
        q.title = request.form['title']
        q.book_id = request.form.get('book_id') or None
        q.week_num = request.form.get('week_num') or None
        q.difficulty = request.form.get('difficulty', 'medium')
        q.tags = request.form.get('tags')
        q.cat_large  = request.form.get('cat_large') or None
        q.cat_medium = request.form.get('cat_medium') or None
        q.cat_small  = request.form.get('cat_small') or None
        q.data = {
            'word': request.form.get('word', ''),
            'definition': request.form.get('definition', ''),
            'choices': choices,
            'correct_idx': int(request.form.get('correct_idx', 0)),
        }
        db.session.commit()
        flash('수정되었습니다.', 'success')
        return redirect(url_for('cms.vocab_list'))
    return render_template('cms/vocab/form.html', question=q, books=_all_books(),
                           difficulty_choices=DIFFICULTY_CHOICES,
                           vocab_categories=VOCAB_CATEGORIES)


@cms_bp.route('/vocab/<question_id>/delete', methods=['POST'])
@login_required
def vocab_delete(question_id):
    if not _hq_only(): abort(403)
    q = BankQuestion.query.filter_by(question_id=question_id, type='vocab_quiz').first_or_404()
    q.is_active = False
    db.session.commit()
    flash('삭제되었습니다.', 'success')
    return redirect(url_for('cms.vocab_list'))


@cms_bp.route('/vocab/template')
@login_required
def vocab_template():
    """어휘 퀴즈 엑셀 템플릿 다운로드"""
    if not _hq_only(): abort(403)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = '어휘퀴즈'

    headers = ['제목', '단어', '뜻(정답)', '보기1', '보기2', '보기3', '보기4',
               '정답번호(1~4)', '난이도(easy/medium/hard)', '주차(숫자)', '태그(쉼표구분)',
               '대분류', '중분류', '소분류']
    col_widths = [25, 15, 25, 20, 20, 20, 20, 16, 22, 12, 25, 22, 20, 18]

    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    req_fill   = PatternFill('solid', fgColor='EEF2FF')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 36

    # 예시 데이터 3행
    samples = [
        ['호기심 많은 소년 - 어휘1', '탐구하다', '무엇인가를 깊이 파고들어 연구하다',
         '탐구하다', '포기하다', '무시하다', '즐기다', 1, 'medium', 1, '독서,어휘',
         '배경지식·스키마 어휘', '문학', '소설'],
        ['분석하다 - 도구어1', '분석하다', '대상을 여러 요소로 나누어 살펴보다',
         '분석하다', '회피하다', '포기하다', '나열하다', 1, 'medium', '', '',
         '학습 도구어', '사고·인지 동사', '분석'],
        ['인과 관계 - 접속어', '따라서', '앞의 내용이 원인이 되어 결론을 이끄는 말',
         '그러나', '반면에', '따라서', '또한', 3, 'easy', '', '',
         '학습 도구어', '접속·연결어', '인과'],
    ]
    for row_num, row_data in enumerate(samples, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = req_fill
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    # 안내 시트
    ws2 = wb.create_sheet('작성 안내')
    notes = [
        ('컬럼', '설명', '필수'),
        ('제목', '문항 분류 이름 (예: 책제목-어휘1)', 'O'),
        ('단어', '테스트할 단어', 'O'),
        ('뜻(정답)', '단어의 올바른 뜻 (정답 보기)', 'O'),
        ('보기1~4', '4개 보기 입력. 정답번호에 해당하는 보기가 뜻과 같아야 함', 'O'),
        ('정답번호', '1~4 중 정답 보기 번호', 'O'),
        ('난이도', 'easy / medium / hard 중 하나 (빈칸이면 medium)', 'X'),
        ('주차', '커리큘럼 주차 숫자 (빈칸 가능)', 'X'),
        ('태그', '쉼표로 구분 (예: 독서,어휘,초등)', 'X'),
        ('대분류', '배경지식·스키마 어휘 또는 학습 도구어', 'X'),
        ('중분류', '대분류에 속하는 중분류 (예: 문학, 사고·인지 동사)', 'X'),
        ('소분류', '중분류에 속하는 소분류 (예: 소설, 분석)', 'X'),
    ]
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 55
    ws2.column_dimensions['C'].width = 8
    for r, (a, b, c) in enumerate(notes, 1):
        ws2.cell(row=r, column=1, value=a).font = Font(bold=(r == 1))
        ws2.cell(row=r, column=2, value=b)
        ws2.cell(row=r, column=3, value=c)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='어휘퀴즈_업로드양식.xlsx')


@cms_bp.route('/vocab/bulk-upload', methods=['POST'])
@login_required
def vocab_bulk_upload():
    """엑셀 파일로 어휘 퀴즈 일괄 등록"""
    if not _hq_only(): abort(403)
    from openpyxl import load_workbook
    from io import BytesIO

    file = request.files.get('excel_file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('xlsx 파일을 선택해주세요.', 'error')
        return redirect(url_for('cms.vocab_list'))

    try:
        wb = load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active
    except Exception:
        flash('파일을 읽을 수 없습니다. 올바른 xlsx 파일인지 확인해주세요.', 'error')
        return redirect(url_for('cms.vocab_list'))

    DIFFICULTY_VALID = {'easy', 'medium', 'hard'}
    saved = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(row):  # 빈 행 스킵
            continue

        title     = str(row[0]).strip()  if row[0] else ''
        word      = str(row[1]).strip()  if row[1] else ''
        definition= str(row[2]).strip()  if row[2] else ''
        choices   = [str(row[i]).strip() if row[i] else '' for i in range(3, 7)]
        correct_raw = row[7]
        difficulty  = str(row[8]).strip().lower() if row[8] else 'medium'
        week_num    = int(row[9]) if row[9] and str(row[9]).strip().isdigit() else None
        tags        = str(row[10]).strip() if row[10] else ''
        cat_large   = str(row[11]).strip() if len(row) > 11 and row[11] else None
        cat_medium  = str(row[12]).strip() if len(row) > 12 and row[12] else None
        cat_small   = str(row[13]).strip() if len(row) > 13 and row[13] else None

        # 필수값 검증
        if not title or not word or not definition:
            errors.append(f'{row_num}행: 제목/단어/뜻은 필수입니다.')
            continue
        if not all(choices):
            errors.append(f'{row_num}행: 보기1~4를 모두 입력해주세요.')
            continue
        try:
            correct_idx = int(correct_raw) - 1
            if correct_idx not in range(4):
                raise ValueError
        except (TypeError, ValueError):
            errors.append(f'{row_num}행: 정답번호는 1~4 사이 숫자여야 합니다.')
            continue
        if difficulty not in DIFFICULTY_VALID:
            difficulty = 'medium'

        q = BankQuestion(
            type='vocab_quiz',
            title=title,
            difficulty=difficulty,
            week_num=week_num,
            tags=tags or None,
            cat_large=cat_large,
            cat_medium=cat_medium,
            cat_small=cat_small,
            data={'word': word, 'definition': definition,
                  'choices': choices, 'correct_idx': correct_idx},
            created_by=current_user.user_id,
        )
        db.session.add(q)
        saved += 1

    if saved:
        db.session.commit()

    if errors:
        flash(f'{saved}개 등록 완료. 오류 {len(errors)}건: ' + ' / '.join(errors[:3])
              + ('...' if len(errors) > 3 else ''), 'warning' if saved else 'error')
    else:
        flash(f'{saved}개 어휘 퀴즈가 등록되었습니다.', 'success')

    return redirect(url_for('cms.vocab_list'))


# ═══════════════════════════════════════════════
# 2. 독서 퀴즈 관리 (간단한 내용 확인 퀴즈)
# ═══════════════════════════════════════════════

@cms_bp.route('/book-quiz')
@login_required
def book_quiz_list():
    if not _hq_only(): abort(403)
    q       = request.args.get('q', '').strip()
    book_id = request.args.get('book_id', '')
    filter_large  = request.args.get('cat_large', '')
    filter_medium = request.args.get('cat_medium', '')
    filter_small  = request.args.get('cat_small', '')
    query = BankQuestion.query.filter_by(type='book_quiz', is_active=True)
    if q:
        query = query.filter(BankQuestion.title.ilike(f'%{q}%'))
    if book_id:
        query = query.filter_by(book_id=book_id)
    if filter_large:
        query = query.filter_by(cat_large=filter_large)
    if filter_medium:
        query = query.filter_by(cat_medium=filter_medium)
    if filter_small:
        query = query.filter_by(cat_small=filter_small)
    questions = query.order_by(BankQuestion.created_at.desc()).all()
    return render_template('cms/book_quiz/list.html', questions=questions,
                           q=q, book_id=book_id, books=_all_books(),
                           filter_large=filter_large, filter_medium=filter_medium,
                           filter_small=filter_small,
                           reading_categories=READING_CATEGORIES,
                           difficulty_choices=DIFFICULTY_CHOICES)


@cms_bp.route('/book-quiz/new', methods=['GET', 'POST'])
@login_required
def book_quiz_new():
    if not _hq_only(): abort(403)
    if request.method == 'POST':
        fmt = request.form.get('format', 'multiple')
        question_text = request.form.get('question', '')
        explanation   = request.form.get('explanation', '')
        if fmt == 'ox':
            data = {'format': 'ox', 'question': question_text,
                    'correct': request.form.get('ox_correct', 'O'), 'explanation': explanation}
        elif fmt == 'short':
            data = {'format': 'short', 'question': question_text,
                    'correct_answer': request.form.get('correct_answer', ''), 'explanation': explanation}
        else:
            choices = [request.form.get(f'choice_{i}', '') for i in range(4)]
            data = {'format': 'multiple', 'question': question_text,
                    'choices': choices, 'correct_idx': int(request.form.get('correct_idx', 0)),
                    'explanation': explanation}
        title = request.form.get('title', '').strip() or question_text[:40]
        bq = BankQuestion(
            type='book_quiz',
            title=title,
            book_id=request.form.get('book_id') or None,
            week_num=request.form.get('week_num') or None,
            difficulty=request.form.get('difficulty', 'medium'),
            cat_large=request.form.get('cat_large') or None,
            cat_medium=request.form.get('cat_medium') or None,
            cat_small=request.form.get('cat_small') or None,
            tags=request.form.get('tags'),
            data=data,
            created_by=current_user.user_id,
        )
        db.session.add(bq)
        db.session.commit()
        flash('독서 퀴즈가 등록되었습니다.', 'success')
        return redirect(url_for('cms.book_quiz_list'))
    return render_template('cms/book_quiz/form.html', books=_all_books(),
                           difficulty_choices=DIFFICULTY_CHOICES,
                           reading_categories=READING_CATEGORIES)


@cms_bp.route('/book-quiz/<question_id>/edit', methods=['GET', 'POST'])
@login_required
def book_quiz_edit(question_id):
    if not _hq_only(): abort(403)
    bq = BankQuestion.query.filter_by(question_id=question_id, type='book_quiz').first_or_404()
    if request.method == 'POST':
        fmt = request.form.get('format', 'multiple')
        question_text = request.form.get('question', '')
        explanation   = request.form.get('explanation', '')
        if fmt == 'ox':
            data = {'format': 'ox', 'question': question_text,
                    'correct': request.form.get('ox_correct', 'O'), 'explanation': explanation}
        elif fmt == 'short':
            data = {'format': 'short', 'question': question_text,
                    'correct_answer': request.form.get('correct_answer', ''), 'explanation': explanation}
        else:
            choices = [request.form.get(f'choice_{i}', '') for i in range(4)]
            data = {'format': 'multiple', 'question': question_text,
                    'choices': choices, 'correct_idx': int(request.form.get('correct_idx', 0)),
                    'explanation': explanation}
        bq.title      = request.form.get('title', '').strip() or bq.title
        bq.book_id    = request.form.get('book_id') or None
        bq.week_num   = request.form.get('week_num') or None
        bq.difficulty = request.form.get('difficulty', 'medium')
        bq.cat_large  = request.form.get('cat_large') or None
        bq.cat_medium = request.form.get('cat_medium') or None
        bq.cat_small  = request.form.get('cat_small') or None
        bq.tags       = request.form.get('tags')
        bq.data       = data
        bq.updated_at = datetime.utcnow()
        db.session.commit()
        flash('독서 퀴즈가 수정되었습니다.', 'success')
        return redirect(url_for('cms.book_quiz_list'))
    return render_template('cms/book_quiz/form.html', question=bq, books=_all_books(),
                           difficulty_choices=DIFFICULTY_CHOICES,
                           reading_categories=READING_CATEGORIES)


@cms_bp.route('/book-quiz/<question_id>/delete', methods=['POST'])
@login_required
def book_quiz_delete(question_id):
    if not _hq_only(): abort(403)
    bq = BankQuestion.query.filter_by(question_id=question_id, type='book_quiz').first_or_404()
    bq.is_active = False
    db.session.commit()
    flash('삭제되었습니다.', 'success')
    return redirect(url_for('cms.book_quiz_list'))


@cms_bp.route('/book-quiz/template')
@login_required
def book_quiz_template():
    """독서 퀴즈 엑셀 템플릿 다운로드"""
    if not _hq_only(): abort(403)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = '독서퀴즈'

    headers = ['제목', '형식(ox/multiple/short)', '문제',
               '보기1', '보기2', '보기3', '보기4',
               '정답(OX→O또는X / 4지선다→1~4 / 단답→직접입력)',
               '해설', '난이도(easy/medium/hard)', '주차(숫자)', '태그(쉼표구분)',
               '대분류', '중분류', '소분류']
    col_widths = [22, 20, 40, 18, 18, 18, 18, 38, 30, 22, 12, 20, 18, 18, 18]

    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    req_fill    = PatternFill('solid', fgColor='EEF2FF')
    thin        = Side(style='thin', color='D1D5DB')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 50

    samples = [
        ['OX 예시 - 홍길동전', 'ox', '홍길동은 서얼 출신이다.',
         '', '', '', '', 'O', '홍길동은 대표적인 서얼 차별 소설의 주인공이다.',
         'medium', 1, '', '문학', '소설·동화', '중1'],
        ['4지선다 예시 - 과학', 'multiple', '광합성이 일어나는 장소는?',
         '미토콘드리아', '엽록체', '핵', '세포막', '2',
         '엽록체에서 빛에너지를 이용해 포도당을 만든다.',
         'easy', 2, '', '비문학', '설명문·정보글', '초5'],
        ['단답 예시 - 역사', 'short', '조선을 건국한 인물은?',
         '', '', '', '', '이성계', '1392년 이성계가 고려를 무너뜨리고 조선을 건국했다.',
         'medium', '', '', '비문학', '사회·역사', '중2'],
    ]
    for row_num, row_data in enumerate(samples, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = req_fill
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[row_num].height = 28

    ws2 = wb.create_sheet('작성 안내')
    notes = [
        ('컬럼', '설명', '필수'),
        ('제목', '문항 분류 이름 (빈칸 시 문제 앞 40자 자동 사용)', 'X'),
        ('형식', 'ox / multiple / short 중 하나', 'O'),
        ('문제', '학생에게 제시할 질문', 'O'),
        ('보기1~4', '4지선다(multiple)일 때만 입력 (나머지는 비워도 됨)', '조건부'),
        ('정답', 'OX→O 또는 X / 4지선다→1~4 숫자 / 단답→정답 텍스트', 'O'),
        ('해설', '정답 해설 (선택)', 'X'),
        ('난이도', 'easy / medium / hard (빈칸이면 medium)', 'X'),
        ('주차', '커리큘럼 주차 숫자 (빈칸 가능)', 'X'),
        ('태그', '쉼표로 구분 (예: 독서,퀴즈)', 'X'),
        ('대분류', '문학 또는 비문학', 'X'),
        ('중분류', '소설·동화, 설명문·정보글 등', 'X'),
        ('소분류', '초1~고등 등 학년', 'X'),
    ]
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 55
    ws2.column_dimensions['C'].width = 8
    for r, (a, b, c) in enumerate(notes, 1):
        ws2.cell(row=r, column=1, value=a).font = Font(bold=(r == 1))
        ws2.cell(row=r, column=2, value=b)
        ws2.cell(row=r, column=3, value=c)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='독서퀴즈_업로드양식.xlsx')


@cms_bp.route('/book-quiz/bulk-upload', methods=['POST'])
@login_required
def book_quiz_bulk_upload():
    """엑셀 파일로 독서 퀴즈 일괄 등록"""
    if not _hq_only(): abort(403)
    from openpyxl import load_workbook
    from io import BytesIO

    file = request.files.get('excel_file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('xlsx 파일을 선택해주세요.', 'error')
        return redirect(url_for('cms.book_quiz_list'))

    try:
        wb = load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active
    except Exception:
        flash('파일을 읽을 수 없습니다. 올바른 xlsx 파일인지 확인해주세요.', 'error')
        return redirect(url_for('cms.book_quiz_list'))

    DIFFICULTY_VALID = {'easy', 'medium', 'hard'}
    FORMATS_VALID    = {'ox', 'multiple', 'short'}
    saved  = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(row):
            continue

        def cell(i): return str(row[i]).strip() if len(row) > i and row[i] is not None else ''

        title      = cell(0)
        fmt        = cell(1).lower()
        question   = cell(2)
        choices    = [cell(3), cell(4), cell(5), cell(6)]
        answer_raw = cell(7)
        explanation= cell(8)
        difficulty = cell(9).lower() or 'medium'
        week_raw   = cell(10)
        tags       = cell(11)
        cat_large  = cell(12) or None
        cat_medium = cell(13) or None
        cat_small  = cell(14) or None

        if fmt not in FORMATS_VALID:
            errors.append(f'{row_num}행: 형식은 ox/multiple/short 중 하나여야 합니다.')
            continue
        if not question:
            errors.append(f'{row_num}행: 문제는 필수입니다.')
            continue
        if not answer_raw:
            errors.append(f'{row_num}행: 정답을 입력해주세요.')
            continue

        if fmt == 'ox':
            correct = answer_raw.upper()
            if correct not in ('O', 'X'):
                errors.append(f'{row_num}행: OX 정답은 O 또는 X여야 합니다.')
                continue
            data = {'format': 'ox', 'question': question,
                    'correct': correct, 'explanation': explanation}
        elif fmt == 'multiple':
            if not all(choices):
                errors.append(f'{row_num}행: 4지선다는 보기1~4를 모두 입력해야 합니다.')
                continue
            try:
                correct_idx = int(answer_raw) - 1
                if correct_idx not in range(4):
                    raise ValueError
            except (TypeError, ValueError):
                errors.append(f'{row_num}행: 4지선다 정답은 1~4 숫자여야 합니다.')
                continue
            data = {'format': 'multiple', 'question': question,
                    'choices': choices, 'correct_idx': correct_idx, 'explanation': explanation}
        else:  # short
            data = {'format': 'short', 'question': question,
                    'correct_answer': answer_raw, 'explanation': explanation}

        if difficulty not in DIFFICULTY_VALID:
            difficulty = 'medium'
        week_num = int(week_raw) if week_raw.isdigit() else None
        auto_title = title or question[:40]

        bq = BankQuestion(
            type='book_quiz',
            title=auto_title,
            difficulty=difficulty,
            week_num=week_num,
            tags=tags or None,
            cat_large=cat_large,
            cat_medium=cat_medium,
            cat_small=cat_small,
            data=data,
            created_by=current_user.user_id,
        )
        db.session.add(bq)
        saved += 1

    if saved:
        db.session.commit()

    if errors:
        flash(f'{saved}개 등록 완료. 오류 {len(errors)}건: ' + ' / '.join(errors[:3])
              + ('...' if len(errors) > 3 else ''), 'warning' if saved else 'error')
    else:
        flash(f'{saved}개 독서 퀴즈가 등록되었습니다.', 'success')

    return redirect(url_for('cms.book_quiz_list'))


# ═══════════════════════════════════════════════
# 3. 토론질문 관리
# ═══════════════════════════════════════════════

@cms_bp.route('/reading-quiz')
@login_required
def reading_quiz_list():
    if not _hq_only(): abort(403)
    q            = request.args.get('q', '').strip()
    book_id      = request.args.get('book_id', '')
    filter_large = request.args.get('cat_large', '')
    filter_medium= request.args.get('cat_medium', '')
    filter_small = request.args.get('cat_small', '')
    filter_rtype = request.args.get('reading_type', '')
    query = BankQuestion.query.filter_by(type='reading_quiz', is_active=True)
    if q:
        query = query.filter(BankQuestion.title.ilike(f'%{q}%'))
    if book_id:
        query = query.filter_by(book_id=book_id)
    if filter_large:
        query = query.filter_by(cat_large=filter_large)
    if filter_medium:
        query = query.filter_by(cat_medium=filter_medium)
    if filter_small:
        query = query.filter_by(cat_small=filter_small)
    if filter_rtype:
        query = query.filter_by(reading_type=filter_rtype)
    questions = query.order_by(BankQuestion.created_at.desc()).all()
    return render_template('cms/reading_quiz/list.html', questions=questions,
                           q=q, book_id=book_id, books=_all_books(),
                           filter_large=filter_large, filter_medium=filter_medium,
                           filter_small=filter_small, filter_rtype=filter_rtype,
                           reading_categories=READING_CATEGORIES,
                           reading_type_choices=READING_TYPE_CHOICES)


@cms_bp.route('/reading-quiz/new', methods=['GET', 'POST'])
@login_required
def reading_quiz_new():
    if not _hq_only(): abort(403)
    if request.method == 'POST':
        data = {
            'step':          request.form.get('step', ''),
            'question_no':   request.form.get('question_no', ''),
            'passage':       request.form.get('passage', ''),
            'page':          request.form.get('page', ''),
            'question':      request.form.get('question', ''),
            'sample_answer': request.form.get('sample_answer', ''),
        }
        step = data['step']
        qno  = data['question_no']
        raw_title = request.form.get('title', '').strip()
        if raw_title:
            title = raw_title
        elif qno:
            title = f'{step} {qno}'.strip() if step else qno
        else:
            title = data['question'][:30] + ('...' if len(data['question']) > 30 else '')

        q = BankQuestion(
            type='reading_quiz',
            title=title,
            book_id=request.form.get('book_id') or None,
            week_num=request.form.get('week_num') or None,
            difficulty=request.form.get('difficulty', 'medium'),
            reading_type=request.form.get('reading_type') or None,
            cat_large=request.form.get('cat_large') or None,
            cat_medium=request.form.get('cat_medium') or None,
            cat_small=request.form.get('cat_small') or None,
            tags=request.form.get('tags'),
            data=data,
            created_by=current_user.user_id,
        )
        db.session.add(q)
        db.session.commit()
        flash('토론질문이 등록되었습니다.', 'success')
        return redirect(url_for('cms.reading_quiz_list'))
    return render_template('cms/reading_quiz/form.html', books=_all_books(),
                           difficulty_choices=DIFFICULTY_CHOICES,
                           reading_categories=READING_CATEGORIES,
                           reading_type_choices=READING_TYPE_CHOICES)


@cms_bp.route('/reading-quiz/<question_id>/edit', methods=['GET', 'POST'])
@login_required
def reading_quiz_edit(question_id):
    if not _hq_only(): abort(403)
    q = BankQuestion.query.filter_by(question_id=question_id, type='reading_quiz').first_or_404()
    if request.method == 'POST':
        q.title       = request.form['title']
        q.book_id     = request.form.get('book_id') or None
        q.week_num    = request.form.get('week_num') or None
        q.difficulty  = request.form.get('difficulty', 'medium')
        q.reading_type= request.form.get('reading_type') or None
        q.cat_large   = request.form.get('cat_large') or None
        q.cat_medium  = request.form.get('cat_medium') or None
        q.cat_small   = request.form.get('cat_small') or None
        q.tags        = request.form.get('tags')
        q.data = {
            'step':          request.form.get('step', ''),
            'question_no':   request.form.get('question_no', ''),
            'passage':       request.form.get('passage', ''),
            'page':          request.form.get('page', ''),
            'question':      request.form.get('question', ''),
            'sample_answer': request.form.get('sample_answer', ''),
        }
        db.session.commit()
        flash('수정되었습니다.', 'success')
        return redirect(url_for('cms.reading_quiz_list'))
    return render_template('cms/reading_quiz/form.html', question=q,
                           books=_all_books(), difficulty_choices=DIFFICULTY_CHOICES,
                           reading_categories=READING_CATEGORIES,
                           reading_type_choices=READING_TYPE_CHOICES)


@cms_bp.route('/reading-quiz/<question_id>/delete', methods=['POST'])
@login_required
def reading_quiz_delete(question_id):
    if not _hq_only(): abort(403)
    q = BankQuestion.query.filter_by(question_id=question_id, type='reading_quiz').first_or_404()
    q.is_active = False
    db.session.commit()
    return redirect(url_for('cms.reading_quiz_list'))


@cms_bp.route('/reading-quiz/template')
@login_required
def reading_quiz_template():
    """독서 퀴즈 엑셀 템플릿 다운로드 (기존 DB 파일 형식 호환)"""
    if not _hq_only(): abort(403)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = '독서논술문항DB'

    # Row 1: 메타 정보 행 (기존 파일 형식과 동일)
    meta_cell = ws.cell(row=1, column=1,
                        value='책제목 | LV레벨 | 분기 주차 | 독서논술 문항 DB')
    meta_cell.font = Font(bold=True, size=11)
    meta_cell.fill = PatternFill('solid', fgColor='1E1B4B')
    meta_cell.font = Font(bold=True, color='FFFFFF', size=11)
    meta_cell.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.row_dimensions[1].height = 28

    # Row 2: 컬럼 헤더 (기존 파일 7개 + 추가 5개)
    headers = ['단계', '문항번호', '독해유형', '지문(발췌문)', '책페이지',
               '질문', '예시답안', '난이도(easy/medium/hard)', '태그(쉼표구분)',
               '대분류', '중분류', '소분류']
    col_widths = [10, 12, 22, 50, 12, 35, 50, 22, 20, 12, 18, 10]

    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    req_fill    = PatternFill('solid', fgColor='EEF2FF')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 36

    # 예시 데이터 (Row 3~4)
    samples = [
        ['2단계', '2-1-1)', '추론적/비판적 독해',
         '유엔 안전보장이사회는 상임 이사국 다섯 나라가 있습니다. 미국, 영국, 프랑스, 러시아, 중국입니다.',
         '120p',
         '유엔 안전보장이사회는 왜 만장일치제를 택했을까요? 다수결을 택했을 경우 일어날 수 있는 혼란에 대해 생각해 봅시다.',
         '다수결 원칙을 따른다면 다섯 나라 중 세 나라만 동의해도 합의가 된 것으로 인정된다...',
         'medium', '정치,독서논술', '비문학', '사회·역사', '중1'],
        ['3단계', '3-1', '사실적 독해',
         '선출직 공무원, 즉 정치인은 목표를 설정하며 그 과정을 감독합니다.',
         '195p',
         '임용직 공무원과 선출직 공무원의 뜻은 무엇인가요?',
         '임용직 공무원은 시험을 통해 뽑는 행정 관료로, 결정된 목표를 효율적으로 집행하는 역할을 한다.',
         'easy', '정치,독서논술', '비문학', '사회·역사', '중2'],
    ]
    for row_num, row_data in enumerate(samples, 3):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = req_fill
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[row_num].height = 60

    # 안내 시트
    ws2 = wb.create_sheet('작성 안내')
    notes = [
        ('컬럼', '설명', '필수'),
        ('(1행)', '책제목/레벨/주차 등 메타 정보 — 업로드 시 자동 스킵됨', '-'),
        ('(2행)', '헤더 행 — 수정하지 마세요', '-'),
        ('단계', '2단계, 3단계 등 레벨 구분', 'X'),
        ('문항번호', '2-1-1), 3-2 등 문항 식별 번호', 'X'),
        ('독해유형', '사실적 / 분석적 / 추론적 / 적용적 / 비판적 독해 (복합 가능: 추론적/비판적 독해)', 'X'),
        ('지문(발췌문)', '책에서 발췌한 지문 텍스트 (없으면 빈칸)', 'X'),
        ('책페이지', '120p, 128~129p 등 출처 페이지', 'X'),
        ('질문', '학생에게 제시할 서술형 문항 질문', 'O'),
        ('예시답안', '모범 답안 / 예시 답안', 'X'),
        ('난이도', 'easy / medium / hard (빈칸이면 medium)', 'X'),
        ('태그', '쉼표로 구분 (예: 독서,사회,중등)', 'X'),
        ('대분류', '문학 또는 비문학', 'X'),
        ('중분류', '설명문·정보글 / 논설문·주장글 / 사회·역사 / 과학·기술 / 소설·동화 등', 'X'),
        ('소분류', '초1 / 초2 / 초3 / 초4 / 초5 / 초6 / 중1 / 중2 / 중3 / 고등', 'X'),
    ]
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 75
    ws2.column_dimensions['C'].width = 8
    for r, (a, b, c) in enumerate(notes, 1):
        ws2.cell(row=r, column=1, value=a).font = Font(bold=(r == 1))
        ws2.cell(row=r, column=2, value=b)
        ws2.cell(row=r, column=3, value=c)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='독서논술문항DB_업로드양식.xlsx')


@cms_bp.route('/reading-quiz/bulk-upload', methods=['POST'])
@login_required
def reading_quiz_bulk_upload():
    """엑셀 파일로 독서 퀴즈 일괄 등록 (기존 DB 파일 형식 지원)
    Row 1: 메타 정보 (스킵)
    Row 2: 헤더 (스킵)
    Row 3~: 데이터
    컬럼: 단계, 문항번호, 독해유형, 지문, 책페이지, 질문, 예시답안, 난이도, 태그, 대분류, 중분류, 소분류
    """
    if not _hq_only(): abort(403)
    from openpyxl import load_workbook
    from io import BytesIO

    file = request.files.get('excel_file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('xlsx 파일을 선택해주세요.', 'error')
        return redirect(url_for('cms.reading_quiz_list'))

    try:
        wb = load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active
    except Exception:
        flash('파일을 읽을 수 없습니다. 올바른 xlsx 파일인지 확인해주세요.', 'error')
        return redirect(url_for('cms.reading_quiz_list'))

    DIFFICULTY_VALID = {'easy', 'medium', 'hard'}
    saved = 0
    errors = []

    # Row 1: 메타, Row 2: 헤더 → Row 3부터 데이터
    for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
        if not any(row):
            continue

        def _s(v): return str(v).strip() if v else ''

        step         = _s(row[0])
        question_no  = _s(row[1])
        reading_type = _s(row[2]) or None
        passage      = _s(row[3])
        page         = _s(row[4])
        question     = _s(row[5])
        sample_answer= _s(row[6])
        difficulty   = _s(row[7]).lower() if len(row) > 7 and row[7] else 'medium'
        tags         = _s(row[8])         if len(row) > 8 else ''
        cat_large    = _s(row[9])  or None if len(row) > 9 else None
        cat_medium   = _s(row[10]) or None if len(row) > 10 else None
        cat_small    = _s(row[11]) or None if len(row) > 11 else None

        if not question:
            errors.append(f'{row_num}행: 질문은 필수입니다.')
            continue
        if difficulty not in DIFFICULTY_VALID:
            difficulty = 'medium'

        # 제목: 문항번호가 있으면 "단계 문항번호", 없으면 질문 앞 30자
        if question_no:
            title = f'{step} {question_no}'.strip() if step else question_no
        else:
            title = question[:30] + ('...' if len(question) > 30 else '')

        q = BankQuestion(
            type='reading_quiz',
            title=title,
            difficulty=difficulty,
            reading_type=reading_type,
            tags=tags or None,
            cat_large=cat_large,
            cat_medium=cat_medium,
            cat_small=cat_small,
            data={
                'step':          step,
                'question_no':   question_no,
                'passage':       passage,
                'page':          page,
                'question':      question,
                'sample_answer': sample_answer,
            },
            created_by=current_user.user_id,
        )
        db.session.add(q)
        saved += 1

    if saved:
        db.session.commit()

    if errors:
        flash(f'{saved}개 등록 완료. 오류 {len(errors)}건: ' + ' / '.join(errors[:3])
              + ('...' if len(errors) > 3 else ''), 'warning' if saved else 'error')
    else:
        flash(f'{saved}개 토론질문이 등록되었습니다.', 'success')

    return redirect(url_for('cms.reading_quiz_list'))


# ═══════════════════════════════════════════════
# 3. 독서 강의 영상 관리
# ═══════════════════════════════════════════════

@cms_bp.route('/videos')
@login_required
def video_list():
    if not _hq_only(): abort(403)
    q            = request.args.get('q', '').strip()
    book_id      = request.args.get('book_id', '')
    filter_large  = request.args.get('cat_large', '')
    filter_medium = request.args.get('cat_medium', '')
    filter_small  = request.args.get('cat_small', '')
    query = LectureVideo.query
    if q:
        query = query.filter(LectureVideo.title.ilike(f'%{q}%'))
    if book_id:
        query = query.filter_by(book_id=book_id)
    if filter_large:
        query = query.filter_by(cat_large=filter_large)
    if filter_medium:
        query = query.filter_by(cat_medium=filter_medium)
    if filter_small:
        query = query.filter_by(cat_small=filter_small)
    videos = query.order_by(LectureVideo.created_at.desc()).all()
    return render_template('cms/video/list.html', videos=videos,
                           q=q, book_id=book_id, books=_all_books(),
                           filter_large=filter_large,
                           filter_medium=filter_medium,
                           filter_small=filter_small,
                           reading_categories=READING_CATEGORIES)


@cms_bp.route('/videos/new', methods=['GET', 'POST'])
@login_required
def video_new():
    if not _hq_only(): abort(403)
    if request.method == 'POST':
        url = request.form.get('url', '').strip()
        v = LectureVideo(
            title=request.form['title'],
            description=request.form.get('description'),
            url=url,
            thumbnail_url=request.form.get('thumbnail_url') or None,
            duration_seconds=request.form.get('duration_seconds') or None,
            book_id=request.form.get('book_id') or None,
            week_num=request.form.get('week_num') or None,
            tags=request.form.get('tags'),
            cat_large=request.form.get('cat_large') or None,
            cat_medium=request.form.get('cat_medium') or None,
            cat_small=request.form.get('cat_small') or None,
            is_published=bool(request.form.get('is_published')),
            created_by=current_user.user_id,
        )
        db.session.add(v)
        db.session.commit()
        flash('강의 영상이 등록되었습니다.', 'success')
        return redirect(url_for('cms.video_list'))
    return render_template('cms/video/form.html', books=_all_books(),
                           reading_categories=READING_CATEGORIES)


@cms_bp.route('/videos/<video_id>/edit', methods=['GET', 'POST'])
@login_required
def video_edit(video_id):
    if not _hq_only(): abort(403)
    v = LectureVideo.query.get_or_404(video_id)
    if request.method == 'POST':
        v.title = request.form['title']
        v.description = request.form.get('description')
        v.url = request.form.get('url', '').strip()
        v.thumbnail_url = request.form.get('thumbnail_url') or None
        v.duration_seconds = request.form.get('duration_seconds') or None
        v.book_id = request.form.get('book_id') or None
        v.week_num = request.form.get('week_num') or None
        v.tags = request.form.get('tags')
        v.cat_large  = request.form.get('cat_large') or None
        v.cat_medium = request.form.get('cat_medium') or None
        v.cat_small  = request.form.get('cat_small') or None
        v.is_published = bool(request.form.get('is_published'))
        db.session.commit()
        flash('수정되었습니다.', 'success')
        return redirect(url_for('cms.video_list'))
    return render_template('cms/video/form.html', video=v, books=_all_books(),
                           reading_categories=READING_CATEGORIES)


@cms_bp.route('/videos/<video_id>/delete', methods=['POST'])
@login_required
def video_delete(video_id):
    if not _hq_only(): abort(403)
    v = LectureVideo.query.get_or_404(video_id)
    db.session.delete(v)
    db.session.commit()
    flash('삭제되었습니다.', 'success')
    return redirect(url_for('cms.video_list'))


@cms_bp.route('/videos/<video_id>/toggle-publish', methods=['POST'])
@login_required
def video_toggle_publish(video_id):
    if not _hq_only(): abort(403)
    v = LectureVideo.query.get_or_404(video_id)
    v.is_published = not v.is_published
    db.session.commit()
    return jsonify({'published': v.is_published})


@cms_bp.route('/videos/template')
@login_required
def video_template():
    """강의 영상 엑셀 양식 다운로드"""
    if not _hq_only(): abort(403)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from flask import send_file
    except ImportError:
        flash('openpyxl 패키지가 필요합니다.', 'error')
        return redirect(url_for('cms.video_list'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '강의영상'

    # Row 1: 메타 정보 (병합)
    ws.merge_cells('A1:I1')
    ws['A1'] = '📹 독서 강의 영상 일괄 등록 양식  |  1행(헤더)는 수정하지 마세요'
    ws['A1'].font = Font(bold=True, size=11)
    ws['A1'].fill = PatternFill('solid', fgColor='4F46E5')
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=11)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Row 2: 헤더
    headers = ['제목*', '유튜브URL*', '설명', '대분류', '중분류', '소분류', '주차', '태그', '공개여부(Y/N)']
    header_fill = PatternFill('solid', fgColor='E0E7FF')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Row 3-4: 샘플 데이터
    samples = [
        ['[소설·동화 초3] 강아지똥 해설 강의', 'https://youtu.be/xxxxxxxxxxxx',
         '강아지똥 줄거리 및 독해 포인트 해설', '문학', '소설·동화', '초3', '1', '동화,초등', 'Y'],
        ['[논설문 초5] 환경 보호 주장하기', 'https://youtu.be/yyyyyyyyyyyy',
         '환경 보호 논설문 구조 분석', '비문학', '논설문·주장글', '초5', '2', '비문학,논설문', 'N'],
    ]
    for row_idx, sample in enumerate(samples, 3):
        for col_idx, val in enumerate(sample, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

    # 컬럼 너비
    col_widths = [35, 40, 30, 12, 18, 8, 6, 20, 14]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # 작성 안내 시트
    ws2 = wb.create_sheet('작성 안내')
    notes = [
        ['항목', '설명'],
        ['제목*', '영상 제목 (필수)'],
        ['유튜브URL*', 'YouTube 영상 URL (필수) — https://youtu.be/... 또는 https://youtube.com/watch?v=...'],
        ['설명', '영상 설명 (선택)'],
        ['대분류', '문학 / 비문학'],
        ['중분류', '소설·동화 / 시·동시 / 수필·일기 / 희곡·시나리오 / 설명문·정보글 / 논설문·주장글 / 전기·인물이야기 / 사회·역사 / 과학·기술 / 예술·문화'],
        ['소분류', '초1 / 초2 / 초3 / 초4 / 초5 / 초6 / 중1 / 중2 / 중3 / 고등'],
        ['주차', '숫자만 입력 (예: 1)'],
        ['태그', '쉼표로 구분 (예: 동화,초등,서사)'],
        ['공개여부(Y/N)', 'Y = 공개, N = 비공개 (기본값: N)'],
    ]
    ws2['A1'].font = Font(bold=True)
    for row_data in notes:
        ws2.append(row_data)
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 75

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='강의영상_일괄등록_양식.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@cms_bp.route('/videos/bulk-upload', methods=['POST'])
@login_required
def video_bulk_upload():
    """강의 영상 엑셀 일괄 등록"""
    if not _hq_only(): abort(403)
    try:
        import openpyxl
    except ImportError:
        flash('openpyxl 패키지가 필요합니다.', 'error')
        return redirect(url_for('cms.video_list'))

    file = request.files.get('excel_file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('xlsx 파일을 선택하세요.', 'error')
        return redirect(url_for('cms.video_list'))

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        flash('파일을 읽을 수 없습니다. xlsx 형식인지 확인하세요.', 'error')
        return redirect(url_for('cms.video_list'))

    def _s(cell):
        return str(cell.value).strip() if cell.value is not None else ''

    count = 0
    errors = []
    rows = list(ws.iter_rows(min_row=3))  # Row 1=메타, Row 2=헤더 skip
    if len(rows) > 200:
        flash('한 번에 최대 200행까지 등록 가능합니다.', 'error')
        return redirect(url_for('cms.video_list'))

    for row_num, row in enumerate(rows, 3):
        title = _s(row[0]) if len(row) > 0 else ''
        url   = _s(row[1]) if len(row) > 1 else ''
        if not title or not url:
            continue  # 빈 행 스킵

        description = _s(row[2]) if len(row) > 2 else ''
        cat_large   = _s(row[3]) or None if len(row) > 3 else None
        cat_medium  = _s(row[4]) or None if len(row) > 4 else None
        cat_small   = _s(row[5]) or None if len(row) > 5 else None
        week_num    = _s(row[6]) if len(row) > 6 else ''
        tags        = _s(row[7]) if len(row) > 7 else ''
        is_pub_str  = _s(row[8]).upper() if len(row) > 8 else ''

        try:
            week_num = int(week_num) if week_num.isdigit() else None
        except Exception:
            week_num = None

        is_published = (is_pub_str == 'Y')

        v = LectureVideo(
            title=title,
            description=description or None,
            url=url,
            cat_large=cat_large,
            cat_medium=cat_medium,
            cat_small=cat_small,
            week_num=week_num,
            tags=tags or None,
            is_published=is_published,
            created_by=current_user.user_id,
        )
        db.session.add(v)
        count += 1

    if count:
        db.session.commit()
        flash(f'{count}개의 강의 영상이 등록되었습니다.', 'success')
    else:
        flash('등록할 데이터가 없습니다. 양식을 확인하세요.', 'warning')
    return redirect(url_for('cms.video_list'))


# ═══════════════════════════════════════════════
# 4. 서술형 문항 관리
# ═══════════════════════════════════════════════

@cms_bp.route('/essay-questions')
@login_required
def essay_question_list():
    if not _hq_only(): abort(403)
    q = request.args.get('q', '').strip()
    book_id = request.args.get('book_id', '')
    query = BankQuestion.query.filter_by(type='essay', is_active=True)
    if q:
        query = query.filter(BankQuestion.title.ilike(f'%{q}%'))
    if book_id:
        query = query.filter_by(book_id=book_id)
    questions = query.order_by(BankQuestion.created_at.desc()).all()
    return render_template('cms/essay_question/list.html', questions=questions,
                           q=q, book_id=book_id, books=_all_books())


@cms_bp.route('/essay-questions/new', methods=['GET', 'POST'])
@login_required
def essay_question_new():
    if not _hq_only(): abort(403)
    if request.method == 'POST':
        data = {
            'question': request.form.get('question', ''),
            'prompt': request.form.get('prompt', ''),
            'rubric': request.form.get('rubric', ''),
            'max_score': float(request.form.get('max_score') or 100),
            'sample_answer': request.form.get('sample_answer', ''),
        }
        q = BankQuestion(
            type='essay',
            title=request.form['title'],
            book_id=request.form.get('book_id') or None,
            week_num=request.form.get('week_num') or None,
            difficulty=request.form.get('difficulty', 'medium'),
            tags=request.form.get('tags'),
            data=data,
            created_by=current_user.user_id,
        )
        db.session.add(q)
        db.session.commit()
        flash('글쓰기 질문이 등록되었습니다.', 'success')
        return redirect(url_for('cms.essay_question_list'))
    return render_template('cms/essay_question/form.html', books=_all_books(),
                           difficulty_choices=DIFFICULTY_CHOICES)


@cms_bp.route('/essay-questions/<question_id>/edit', methods=['GET', 'POST'])
@login_required
def essay_question_edit(question_id):
    if not _hq_only(): abort(403)
    q = BankQuestion.query.filter_by(question_id=question_id, type='essay').first_or_404()
    if request.method == 'POST':
        q.title = request.form['title']
        q.book_id = request.form.get('book_id') or None
        q.week_num = request.form.get('week_num') or None
        q.difficulty = request.form.get('difficulty', 'medium')
        q.tags = request.form.get('tags')
        q.data = {
            'question': request.form.get('question', ''),
            'prompt': request.form.get('prompt', ''),
            'rubric': request.form.get('rubric', ''),
            'max_score': float(request.form.get('max_score') or 100),
            'sample_answer': request.form.get('sample_answer', ''),
        }
        db.session.commit()
        flash('수정되었습니다.', 'success')
        return redirect(url_for('cms.essay_question_list'))
    return render_template('cms/essay_question/form.html', question=q,
                           books=_all_books(), difficulty_choices=DIFFICULTY_CHOICES)


@cms_bp.route('/essay-questions/<question_id>/delete', methods=['POST'])
@login_required
def essay_question_delete(question_id):
    if not _hq_only(): abort(403)
    q = BankQuestion.query.filter_by(question_id=question_id, type='essay').first_or_404()
    q.is_active = False
    db.session.commit()
    return redirect(url_for('cms.essay_question_list'))


@cms_bp.route('/essay-questions/template')
@login_required
def essay_question_template():
    """글쓰기 질문 엑셀 템플릿 다운로드"""
    if not _hq_only(): abort(403)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = '글쓰기질문'

    headers = ['제목', '문제 질문', '문제 지문(선택)', '채점 기준(루브릭)',
               '예시 답안', '만점(숫자)', '난이도(easy/medium/hard)', '주차(숫자)', '태그(쉼표구분)']
    col_widths = [25, 45, 45, 45, 45, 12, 22, 12, 22]

    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    req_fill    = PatternFill('solid', fgColor='EEF2FF')
    thin        = Side(style='thin', color='D1D5DB')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 36

    samples = [
        ['홍길동전 - 주제 파악', '홍길동이 집을 떠난 이유를 자신의 말로 설명해보세요.',
         '홍길동은 서얼 출신으로 능력이 있음에도 불구하고 신분 차별을 받았다.',
         'A(90~100): 신분 차별과 자아실현 욕구를 모두 언급\nB(70~89): 한 가지만 언급',
         '홍길동은 서얼 차별로 인해 뜻을 펼칠 수 없었기 때문에 집을 떠났습니다.',
         100, 'medium', 1, '독서,논술'],
        ['과학 글쓰기 - 광합성', '광합성이 일어나는 과정을 단계별로 설명하세요.',
         '', '', '', 100, 'hard', 3, '과학,서술'],
    ]
    for row_num, row_data in enumerate(samples, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = req_fill
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[row_num].height = 40

    ws2 = wb.create_sheet('작성 안내')
    notes = [
        ('컬럼', '설명', '필수'),
        ('제목', '문항 분류 이름', 'O'),
        ('문제 질문', '학생에게 제시하는 핵심 질문', 'O'),
        ('문제 지문', '참고 지문 또는 배경 설명 (없으면 비워도 됨)', 'X'),
        ('채점 기준', '루브릭 (없으면 비워도 됨)', 'X'),
        ('예시 답안', '모범 답안 (없으면 비워도 됨)', 'X'),
        ('만점', '숫자 (기본값 100)', 'X'),
        ('난이도', 'easy / medium / hard (기본값 medium)', 'X'),
        ('주차', '커리큘럼 주차 숫자', 'X'),
        ('태그', '쉼표로 구분', 'X'),
    ]
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 8
    for r, (a, b, c) in enumerate(notes, 1):
        ws2.cell(row=r, column=1, value=a).font = Font(bold=(r == 1))
        ws2.cell(row=r, column=2, value=b)
        ws2.cell(row=r, column=3, value=c)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='글쓰기질문_업로드양식.xlsx')


@cms_bp.route('/essay-questions/bulk-upload', methods=['POST'])
@login_required
def essay_question_bulk_upload():
    """엑셀 파일로 글쓰기 질문 일괄 등록"""
    if not _hq_only(): abort(403)
    from openpyxl import load_workbook
    from io import BytesIO

    file = request.files.get('excel_file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('xlsx 파일을 선택해주세요.', 'error')
        return redirect(url_for('cms.essay_question_list'))

    try:
        wb = load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active
    except Exception:
        flash('파일을 읽을 수 없습니다. 올바른 xlsx 파일인지 확인해주세요.', 'error')
        return redirect(url_for('cms.essay_question_list'))

    DIFFICULTY_VALID = {'easy', 'medium', 'hard'}
    saved  = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(row):
            continue

        def cell(i): return str(row[i]).strip() if len(row) > i and row[i] is not None else ''

        title         = cell(0)
        question      = cell(1)
        prompt        = cell(2)
        rubric        = cell(3)
        sample_answer = cell(4)
        max_score_raw = cell(5)
        difficulty    = cell(6).lower() or 'medium'
        week_raw      = cell(7)
        tags          = cell(8)

        if not title:
            errors.append(f'{row_num}행: 제목은 필수입니다.')
            continue
        if not question:
            errors.append(f'{row_num}행: 문제 질문은 필수입니다.')
            continue

        try:
            max_score = float(max_score_raw) if max_score_raw else 100.0
        except ValueError:
            max_score = 100.0
        if difficulty not in DIFFICULTY_VALID:
            difficulty = 'medium'
        week_num = int(week_raw) if week_raw.isdigit() else None

        q = BankQuestion(
            type='essay',
            title=title,
            difficulty=difficulty,
            week_num=week_num,
            tags=tags or None,
            data={'question': question, 'prompt': prompt, 'rubric': rubric,
                  'max_score': max_score, 'sample_answer': sample_answer},
            created_by=current_user.user_id,
        )
        db.session.add(q)
        saved += 1

    if saved:
        db.session.commit()

    if errors:
        flash(f'{saved}개 등록 완료. 오류 {len(errors)}건: ' + ' / '.join(errors[:3])
              + ('...' if len(errors) > 3 else ''), 'warning' if saved else 'error')
    else:
        flash(f'{saved}개 글쓰기 질문이 등록되었습니다.', 'success')

    return redirect(url_for('cms.essay_question_list'))


# ═══════════════════════════════════════════════
# 5. 모의고사 관리
# ═══════════════════════════════════════════════

@cms_bp.route('/mock-exams')
@login_required
def mock_exam_list():
    if not _hq_only(): abort(403)
    exams = MockExam.query.order_by(MockExam.created_at.desc()).all()
    return render_template('cms/exam/list.html', exams=exams)


@cms_bp.route('/mock-exams/new', methods=['GET', 'POST'])
@login_required
def mock_exam_new():
    if not _hq_only(): abort(403)
    if request.method == 'POST':
        exam = MockExam(
            title=request.form['title'],
            description=request.form.get('description'),
            time_limit_minutes=request.form.get('time_limit_minutes') or None,
            book_id=request.form.get('book_id') or None,
            week_num=request.form.get('week_num') or None,
            tags=request.form.get('tags'),
            is_published=bool(request.form.get('is_published')),
            created_by=current_user.user_id,
        )
        db.session.add(exam)
        db.session.commit()
        flash('모의고사가 생성되었습니다. 문항을 추가하세요.', 'success')
        return redirect(url_for('cms.mock_exam_detail', exam_id=exam.exam_id))
    return render_template('cms/exam/form.html', books=_all_books())


@cms_bp.route('/mock-exams/<exam_id>')
@login_required
def mock_exam_detail(exam_id):
    if not _hq_only(): abort(403)
    exam = MockExam.query.get_or_404(exam_id)
    return render_template('cms/exam/detail.html', exam=exam,
                           question_types=EXAM_QUESTION_TYPES)


@cms_bp.route('/mock-exams/<exam_id>/edit', methods=['GET', 'POST'])
@login_required
def mock_exam_edit(exam_id):
    if not _hq_only(): abort(403)
    exam = MockExam.query.get_or_404(exam_id)
    if request.method == 'POST':
        exam.title = request.form['title']
        exam.description = request.form.get('description')
        exam.time_limit_minutes = request.form.get('time_limit_minutes') or None
        exam.book_id = request.form.get('book_id') or None
        exam.week_num = request.form.get('week_num') or None
        exam.tags = request.form.get('tags')
        exam.is_published = bool(request.form.get('is_published'))
        db.session.commit()
        flash('수정되었습니다.', 'success')
        return redirect(url_for('cms.mock_exam_detail', exam_id=exam_id))
    return render_template('cms/exam/form.html', exam=exam, books=_all_books())


@cms_bp.route('/mock-exams/<exam_id>/delete', methods=['POST'])
@login_required
def mock_exam_delete(exam_id):
    if not _hq_only(): abort(403)
    exam = MockExam.query.get_or_404(exam_id)
    db.session.delete(exam)
    db.session.commit()
    flash('삭제되었습니다.', 'success')
    return redirect(url_for('cms.mock_exam_list'))


# 문항 개별 추가
@cms_bp.route('/mock-exams/<exam_id>/questions/add', methods=['POST'])
@login_required
def mock_exam_add_question(exam_id):
    if not _hq_only(): abort(403)
    exam = MockExam.query.get_or_404(exam_id)
    qtype = request.form.get('question_type', 'multiple_choice')
    choices = None
    if qtype == 'multiple_choice':
        choices = [request.form.get(f'choice_{i}', '') for i in range(4)]
    mq = MockExamQuestion(
        exam_id=exam_id,
        question_type=qtype,
        passage=request.form.get('passage') or None,
        question_text=request.form['question_text'],
        choices=choices,
        correct_answer=request.form.get('correct_answer'),
        explanation=request.form.get('explanation'),
        score=float(request.form.get('score') or 1),
        order_num=len(exam.questions),
    )
    db.session.add(mq)
    db.session.commit()
    flash('문항이 추가되었습니다.', 'success')
    return redirect(url_for('cms.mock_exam_detail', exam_id=exam_id))


# 문항 삭제
@cms_bp.route('/mock-exams/questions/<int:mq_id>/delete', methods=['POST'])
@login_required
def mock_exam_delete_question(mq_id):
    if not _hq_only(): abort(403)
    mq = MockExamQuestion.query.get_or_404(mq_id)
    exam_id = mq.exam_id
    db.session.delete(mq)
    db.session.commit()
    return redirect(url_for('cms.mock_exam_detail', exam_id=exam_id))


# 엑셀 업로드
@cms_bp.route('/mock-exams/<exam_id>/import', methods=['GET', 'POST'])
@login_required
def mock_exam_import(exam_id):
    if not _hq_only(): abort(403)
    exam = MockExam.query.get_or_404(exam_id)
    preview = None
    errors = []

    if request.method == 'POST':
        action = request.form.get('action', 'preview')
        file = request.files.get('excel_file')

        if action == 'preview' and file and file.filename:
            try:
                rows = _parse_exam_excel(file)
                preview = rows
                if not rows:
                    errors.append('파싱된 문항이 없습니다. 양식을 확인해주세요.')
            except Exception as e:
                errors.append(f'파일 오류: {str(e)}')

        elif action == 'save':
            # 저장: form에서 JSON으로 전달된 rows 처리
            import json
            rows_json = request.form.get('rows_json', '[]')
            try:
                rows = json.loads(rows_json)
                start_order = len(exam.questions)
                for i, row in enumerate(rows):
                    mq = MockExamQuestion(
                        exam_id=exam_id,
                        question_type=row.get('type', 'multiple_choice'),
                        passage=row.get('passage') or None,
                        question_text=row['question'],
                        choices=row.get('choices'),
                        correct_answer=str(row.get('correct', '')),
                        explanation=row.get('explanation'),
                        score=float(row.get('score') or 1),
                        order_num=start_order + i,
                    )
                    db.session.add(mq)
                db.session.commit()
                flash(f'{len(rows)}개 문항이 추가되었습니다.', 'success')
                return redirect(url_for('cms.mock_exam_detail', exam_id=exam_id))
            except Exception as e:
                errors.append(f'저장 오류: {str(e)}')

    return render_template('cms/exam/import.html', exam=exam,
                           preview=preview, errors=errors)


def _parse_exam_excel(file_obj):
    """
    엑셀 컬럼 형식:
    A: 문항유형 (객관식/단답형/서술형)
    B: 지문 (optional)
    C: 문제
    D~G: 보기1~4 (객관식만)
    H: 정답
    I: 배점
    J: 해설 (optional)
    """
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(file_obj.read()), read_only=True)
    ws = wb.active
    rows = []
    TYPE_MAP = {'객관식': 'multiple_choice', '단답형': 'short_answer', '서술형': 'essay'}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:  # 문제(C열) 없으면 스킵
            continue
        qtype_raw = str(row[0] or '객관식').strip()
        qtype = TYPE_MAP.get(qtype_raw, 'multiple_choice')
        choices = None
        if qtype == 'multiple_choice':
            choices = [str(row[j] or '') for j in range(3, 7)]
            choices = [c for c in choices if c]

        rows.append({
            'type': qtype,
            'type_display': qtype_raw,
            'passage': str(row[1] or '') or None,
            'question': str(row[2] or ''),
            'choices': choices,
            'correct': str(row[7] or '') if row[7] is not None else '',
            'score': float(row[8] or 1) if row[8] is not None else 1,
            'explanation': str(row[9] or '') if len(row) > 9 and row[9] else '',
        })
    return rows


# ═══════════════════════════════════════════════
# 6. 학습 교재 관리
# ═══════════════════════════════════════════════

@cms_bp.route('/materials')
@login_required
def material_list():
    if not _hq_only(): abort(403)
    q = request.args.get('q', '').strip()
    book_id = request.args.get('book_id', '')
    query = StudyMaterial.query
    if q:
        query = query.filter(StudyMaterial.title.ilike(f'%{q}%'))
    if book_id:
        query = query.filter_by(book_id=book_id)
    materials = query.order_by(StudyMaterial.created_at.desc()).all()
    return render_template('cms/material/list.html', materials=materials,
                           q=q, book_id=book_id, books=_all_books())


@cms_bp.route('/materials/new', methods=['GET', 'POST'])
@login_required
def material_new():
    if not _hq_only(): abort(403)
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('파일을 선택해주세요.', 'warning')
            return render_template('cms/material/form.html', books=_all_books())
        if not _allowed_file(file.filename):
            flash('허용되지 않는 파일 형식입니다.', 'warning')
            return render_template('cms/material/form.html', books=_all_books())

        file_path, file_name, file_type, file_size = _save_file(file)
        m = StudyMaterial(
            title=request.form['title'],
            description=request.form.get('description'),
            file_name=file_name,
            file_path=file_path,
            file_type=file_type,
            file_size=file_size,
            book_id=request.form.get('book_id') or None,
            week_num=request.form.get('week_num') or None,
            tags=request.form.get('tags'),
            is_published=bool(request.form.get('is_published')),
            created_by=current_user.user_id,
        )
        db.session.add(m)
        db.session.commit()
        flash('교재가 등록되었습니다.', 'success')
        return redirect(url_for('cms.material_list'))
    return render_template('cms/material/form.html', books=_all_books())


@cms_bp.route('/materials/<material_id>/edit', methods=['GET', 'POST'])
@login_required
def material_edit(material_id):
    if not _hq_only(): abort(403)
    m = StudyMaterial.query.get_or_404(material_id)
    if request.method == 'POST':
        m.title = request.form['title']
        m.description = request.form.get('description')
        m.book_id = request.form.get('book_id') or None
        m.week_num = request.form.get('week_num') or None
        m.tags = request.form.get('tags')
        m.is_published = bool(request.form.get('is_published'))
        # 파일 교체 (선택)
        file = request.files.get('file')
        if file and file.filename and _allowed_file(file.filename):
            file_path, file_name, file_type, file_size = _save_file(file)
            m.file_path = file_path
            m.file_name = file_name
            m.file_type = file_type
            m.file_size = file_size
        db.session.commit()
        flash('수정되었습니다.', 'success')
        return redirect(url_for('cms.material_list'))
    return render_template('cms/material/form.html', material=m, books=_all_books())


@cms_bp.route('/materials/<material_id>/delete', methods=['POST'])
@login_required
def material_delete(material_id):
    if not _hq_only(): abort(403)
    m = StudyMaterial.query.get_or_404(material_id)
    db.session.delete(m)
    db.session.commit()
    flash('삭제되었습니다.', 'success')
    return redirect(url_for('cms.material_list'))


@cms_bp.route('/materials/<material_id>/download')
@login_required
def material_download(material_id):
    m = StudyMaterial.query.get_or_404(material_id)
    if not m.is_published and not _hq_only():
        abort(403)
    m.download_count += 1
    db.session.commit()
    static_dir = os.path.join(os.path.dirname(__file__), '..', 'static')
    return send_from_directory(static_dir, m.file_path,
                               as_attachment=True,
                               download_name=m.file_name)


@cms_bp.route('/materials/<material_id>/toggle-publish', methods=['POST'])
@login_required
def material_toggle_publish(material_id):
    if not _hq_only(): abort(403)
    m = StudyMaterial.query.get_or_404(material_id)
    m.is_published = not m.is_published
    db.session.commit()
    return jsonify({'published': m.is_published})
